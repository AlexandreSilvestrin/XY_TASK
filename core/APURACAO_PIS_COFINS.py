import os
import re
import unicodedata

import numpy as np
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook

from models.dados_apuracao import DadosApuracaoModel

LOG_MODULE = "apuracao-pis-cofins"


def extrarir_cnpj(texto):
    match = re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", texto)
    if not match:
        raise ValueError("CNPJ não encontrado no arquivo.")
    return re.sub(r"\D", "", match.group())


def extrair_nome(df):
    nome = df.iloc[0, 0]
    nome = "".join(
        char
        for char in unicodedata.normalize("NFD", nome)
        if unicodedata.category(char) != "Mn"
    )
    nome = nome.upper()
    nome = nome.replace("-", " ")
    nome = nome.replace(".", " ")
    nome = nome.replace("/", " ")
    nome = re.sub(r"\s+", " ", nome)
    nome = nome.replace("CONSORCIO", "")
    return nome.strip()


def listar_pasta(pasta):
    return [
        nome
        for nome in os.listdir(pasta)
        if os.path.isfile(os.path.join(pasta, nome))
        and nome.lower().endswith((".xlsx", ".xls"))
    ]


def ver_linha_coluna(tabela):
    wb = load_workbook(tabela, data_only=True)
    ws = wb.active

    texto_busca = "APROPRIAÇÃO DOS CRÉDITOS POR CONSORCIADAS"

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == texto_busca:
                return cell.row - 1, cell.column - 1

    raise ValueError(
        f"Texto '{texto_busca}' não encontrado em {os.path.basename(tabela)}."
    )


def processar_df(df, linha, coluna):
    bloco = df.iloc[linha:, coluna:].copy().reset_index(drop=True)
    df = bloco.drop(0)
    df = df.loc[:, df.iloc[0] != "BASE CALCULO"]
    contador = defaultdict(int)
    novas_colunas = []

    for nome in df.iloc[0]:
        contador[nome] += 1
        novas_colunas.append(f"{nome}_{contador[nome]}")

    df.columns = novas_colunas
    df = df.iloc[1:].reset_index(drop=True)
    df = df.replace(r"^\s*$", np.nan, regex=True)
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")

    return df.iloc[-1].to_dict()


def _valor_para_centavos(valor):
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        raise ValueError("Valor PIS/COFINS ausente (NaN) na planilha.")
    try:
        numero = float(valor)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Valor PIS/COFINS inválido: {valor!r}.") from exc
    if np.isnan(numero):
        raise ValueError("Valor PIS/COFINS ausente (NaN) na planilha.")
    return str(int(round(numero * 100)))


def gerar_df(nome, lista, data):
    linhas = []
    tipos_ordem = ["PIS", "COFINS"]
    nome_arquivo = f"CONSORCIO {nome} -- CREDITO PIS E COFINS"

    for tipo in tipos_ordem:
        for empresa in lista:
            codigos = empresa["codigos"]
            valor_e = _valor_para_centavos(empresa[f"{tipo}_valor"])
            linhas.append(
                {
                    "A": "",
                    "B": codigos[f"{tipo}_B"],
                    "C": codigos[f"{tipo}_C"],
                    "D": 1,
                    "E": valor_e,
                    "F": data,
                    "G": "",
                    "H": (
                        f"CREDITO DO {tipo} REF. NOTAS DE FORNECEDORES DO "
                        f'CONSORCIO NO PERIODO - {empresa["empresa"]}'
                    ),
                }
            )

        linhas.append(
            {
                "A": "",
                "B": "",
                "C": "",
                "D": "",
                "E": "",
                "F": "",
                "G": "",
                "H": "",
            }
        )

    return pd.DataFrame(linhas), nome_arquivo


def organizar_dados(arquivos, pasta):
    lista_dict = []
    erros = []

    for arq in arquivos:
        caminho = os.path.join(pasta, arq)
        try:
            linha, coluna = ver_linha_coluna(caminho)
            df = pd.read_excel(caminho)
            nome = extrair_nome(df)
            dict_pis_cofins = processar_df(df, linha, coluna)
            lista_dict.append((nome, dict_pis_cofins, arq))
        except Exception as exc:
            erros.append((arq, str(exc)))

    return lista_dict, erros


def _resolver_consorcio(codigos, consorcio):
    if consorcio in codigos:
        return consorcio

    target = DadosApuracaoModel.normalize_key(consorcio)
    for key in codigos:
        if DadosApuracaoModel.normalize_key(key) == target:
            return key
    return None


def organizar_arquivo_final(lista_dict, codigos):
    listafinal = []
    erros = []

    for item in lista_dict:
        if len(item) == 3:
            consorcio, pis_cofins, arquivo = item
        else:
            consorcio, pis_cofins = item
            arquivo = consorcio

        try:
            consorcio_key = _resolver_consorcio(codigos, consorcio)
            if not consorcio_key:
                raise ValueError(
                    f"Códigos não encontrados para o consórcio '{consorcio}'."
                )

            codigos_consorcio = codigos[consorcio_key]
            lista_consorcio = []

            for index, empresa in enumerate(codigos_consorcio, start=1):
                chave_pis = f"PIS_{index}"
                chave_cofins = f"COFINS_{index}"
                if chave_pis not in pis_cofins or chave_cofins not in pis_cofins:
                    raise ValueError(
                        f"Planilha não tem PIS/COFINS na posição {index} "
                        f"para '{empresa}'."
                    )

                lista_consorcio.append(
                    {
                        "empresa": empresa,
                        "codigos": codigos_consorcio[empresa],
                        "PIS_valor": pis_cofins[chave_pis],
                        "COFINS_valor": pis_cofins[chave_cofins],
                    }
                )

            listafinal.append((consorcio, lista_consorcio, arquivo))
        except Exception as exc:
            erros.append((arquivo, str(exc)))

    return listafinal, erros


class ApuracaoPisCofinsWeb:
    def __init__(
        self,
        entrada,
        saida,
        empresa,
        data,
        emit_log,
        log_module=LOG_MODULE,
    ):
        self.entrada = entrada
        self.saida = saida
        self.empresa = empresa
        self.data = data
        self.emit_log = emit_log
        self.log_module = log_module

    def executar(self):
        arquivos = listar_pasta(self.entrada)
        if not arquivos:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=self.entrada,
                message="Nenhum arquivo Excel encontrado na pasta de entrada.",
            )
            return False

        os.makedirs(self.saida, exist_ok=True)

        try:
            codigos = DadosApuracaoModel.get_for_processing(self.empresa)
        except Exception as exc:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=self.entrada,
                message=str(exc),
            )
            return False

        if not codigos:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=self.entrada,
                message=f"Nenhum código cadastrado para a empresa '{self.empresa}'.",
            )
            return False

        lista_dict, erros_leitura = organizar_dados(arquivos, self.entrada)
        for arquivo, mensagem in erros_leitura:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=arquivo,
                message=mensagem,
            )

        listafinal, erros_codigos = organizar_arquivo_final(lista_dict, codigos)
        for arquivo, mensagem in erros_codigos:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=arquivo,
                message=mensagem,
            )

        gerados = 0
        erros_geracao = 0

        for consorcio, lista, arquivo in listafinal:
            try:
                df, nome = gerar_df(consorcio, lista, self.data)
                caminho = os.path.join(self.saida, f"{nome}.xlsx")
                df.to_excel(caminho, index=False, header=None)
                gerados += 1
                self.emit_log(
                    module=self.log_module,
                    status="success",
                    file=os.path.basename(caminho),
                    message=f"Apuração gerada para {consorcio}.",
                )
            except Exception as exc:
                erros_geracao += 1
                self.emit_log(
                    module=self.log_module,
                    status="error",
                    file=arquivo,
                    message=f"Erro ao gerar apuração de {consorcio}: {exc}",
                )

        erros_total = len(erros_leitura) + len(erros_codigos) + erros_geracao

        if gerados == 0:
            self.emit_log(
                module=self.log_module,
                status="error",
                file=self.entrada,
                message=(
                    f"Nenhuma apuração foi gerada "
                    f"({erros_total} erro(s) em {len(arquivos)} arquivo(s))."
                ),
            )
            return False

        if erros_total > 0:
            self.emit_log(
                module=self.log_module,
                status="success",
                file=self.entrada,
                message=(
                    f"Apuração parcialmente concluída para {self.empresa}: "
                    f"{gerados} gerado(s), {erros_total} com erro."
                ),
            )
        else:
            self.emit_log(
                module=self.log_module,
                status="success",
                file=self.entrada,
                message=(
                    f"Apuração PIS/COFINS concluída ({gerados} arquivo(s)) "
                    f"para {self.empresa}."
                ),
            )
        return True
